import streamlit as st
from openpyxl import load_workbook
from io import BytesIO
import base64
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

# --- Funciones para el llenado del Excel ---

def diligenciar_hoja_empalmeria(libro, data_general, empalmes_data):
    """Diligencia la hoja 'REPORTE EMPALMERIA' con la información general y los datos de los empalmes."""
    hoja = libro['REPORTE EMPALMERIA']

    hoja['B5'] = data_general.get('cliente', '')
    hoja['I5'] = data_general.get('direccion_cliente', '')
    hoja['B7'] = data_general.get('empalmador', '')
    hoja['I9'] = data_general.get('ot', '')
    hoja['B31'] = data_general.get('fecha', '')
    hoja['C37'] = data_general.get('observaciones', '')

    fila_inicio_empalme = 15
    for i, empalme in enumerate(empalmes_data):
        fila_actual = fila_inicio_empalme + i
        hoja[f'A{fila_actual}'] = empalme.get('num_empalme', '')
        hoja[f'B{fila_actual}'] = empalme.get('tipo', '')
        hoja[f'C{fila_actual}'] = empalme.get('direccion_empalme', '')
        hoja[f'D{fila_actual}'] = empalme.get('num_inventario_in', '')
        hoja[f'E{fila_actual}'] = empalme.get('marca_in', '')
        hoja[f'F{fila_actual}'] = empalme.get('capacidad_in', '')
        hoja[f'G{fila_actual}'] = empalme.get('hilos_in', '')
        hoja[f'H{fila_actual}'] = empalme.get('num_inventario_out', '')
        hoja[f'I{fila_actual}'] = empalme.get('marca_out', '')
        hoja[f'J{fila_actual}'] = empalme.get('capacidad_out', '')
        hoja[f'K{fila_actual}'] = empalme.get('hilos_out', '')
        hoja[f'L{fila_actual}'] = empalme.get('distancia', '')

def cargar_listado_material_con_openpyxl(ruta_excel):
    """Carga los datos de la hoja 'LISTADO MATERIAL' usando openpyxl."""
    try:
        libro = load_workbook(ruta_excel, data_only=True)
        hoja_material = libro['LISTADO MATERIAL']
        data = []
        for row in hoja_material.iter_rows(min_row=3, values_only=True):
            if row[0] is not None:
                data.append({
                    'Numero': row[0],
                    'Descripcion': row[1],
                    'Unidad': row[2],
                    'Valor Unitario': row[3]
                })
        df = pd.DataFrame(data)
        df['Numero'] = pd.to_numeric(df['Numero'], errors='coerce')
        return df
    except Exception as e:
        st.error(f"Error al cargar la hoja 'LISTADO MATERIAL' con openpyxl: {e}")
        return None

def diligenciar_hoja_material_consumido(libro, materiales_instalados):
    """Crea y diligencia la hoja 'MATERIAL CONSUMIDO' con los materiales por categoría."""
    
    # Crear la nueva hoja, si ya existe la sobreescribe
    if 'MATERIAL CONSUMIDO' in libro.sheetnames:
        del libro['MATERIAL CONSUMIDO']
    hoja = libro.create_sheet('MATERIAL CONSUMIDO')
    
    fila_actual = 3
    
    # Material Interno
    hoja[f'A{fila_actual}'] = 'MATERIAL INTERNO'
    hoja[f'A{fila_actual+1}'] = 'NO'
    hoja[f'B{fila_actual+1}'] = 'DESCRIPCIÓN'
    hoja[f'C{fila_actual+1}'] = 'UNIDAD'
    hoja[f'D{fila_actual+1}'] = 'V/U'
    hoja[f'E{fila_actual+1}'] = 'CANTIDAD'
    fila_actual += 2
    for item in [m for m in materiales_instalados if m['Tipo'] == 'MATERIAL INTERNO']:
        hoja[f'A{fila_actual}'] = item['Numero']
        hoja[f'B{fila_actual}'] = item['Descripcion']
        hoja[f'C{fila_actual}'] = item['Unidad']
        hoja[f'D{fila_actual}'] = item['Valor Unitario']
        hoja[f'E{fila_actual}'] = item['Cantidad']
        fila_actual += 1
    
    fila_actual += 2

    # Material Externo
    hoja[f'A{fila_actual}'] = 'MATERIAL EXTERNO'
    hoja[f'A{fila_actual+1}'] = 'NO'
    hoja[f'B{fila_actual+1}'] = 'DESCRIPCIÓN'
    hoja[f'C{fila_actual+1}'] = 'UNIDAD'
    hoja[f'D{fila_actual+1}'] = 'V/U'
    hoja[f'E{fila_actual+1}'] = 'CANTIDAD'
    fila_actual += 2
    for item in [m for m in materiales_instalados if m['Tipo'] == 'MATERIAL EXTERNO']:
        hoja[f'A{fila_actual}'] = item['Numero']
        hoja[f'B{fila_actual}'] = item['Descripcion']
        hoja[f'C{fila_actual}'] = item['Unidad']
        hoja[f'D{fila_actual}'] = item['Valor Unitario']
        hoja[f'E{fila_actual}'] = item['Cantidad']
        fila_actual += 1

    fila_actual += 2

    # Material Empalmería
    hoja[f'A{fila_actual}'] = 'MATERIAL EMPALMERIA'
    hoja[f'A{fila_actual+1}'] = 'NO'
    hoja[f'B{fila_actual+1}'] = 'DESCRIPCIÓN'
    hoja[f'C{fila_actual+1}'] = 'UNIDAD'
    hoja[f'D{fila_actual+1}'] = 'V/U'
    hoja[f'E{fila_actual+1}'] = 'CANTIDAD'
    fila_actual += 2
    for item in [m for m in materiales_instalados if m['Tipo'] == 'MATERIAL EMPALMERIA']:
        hoja[f'A{fila_actual}'] = item['Numero']
        hoja[f'B{fila_actual}'] = item['Descripcion']
        hoja[f'C{fila_actual}'] = item['Unidad']
        hoja[f'D{fila_actual}'] = item['Valor Unitario']
        hoja[f'E{fila_actual}'] = item['Cantidad']
        fila_actual += 1

# --- Lógica de la interfaz de usuario en Streamlit ---

st.set_page_config(layout="wide")
st.title("Formulario de Empalmería y Materiales")
ruta_excel = 'FORMATO1.xlsx'

# Inicializar estados de sesión
if 'general_data' not in st.session_state:
    st.session_state.general_data = {}
if 'empalmes_data' not in st.session_state:
    st.session_state.empalmes_data = []
if 'materiales_consumidos' not in st.session_state:
    st.session_state.materiales_consumidos = []
if 'editando_empalme_idx' not in st.session_state:
    st.session_state.editando_empalme_idx = None

# Selector de hoja principal
hojas_excel = ["REPORTE EMPALMERIA", "MATERIAL CONSUMIDO"]
hoja_seleccionada = st.selectbox("Selecciona la hoja de Excel:", hojas_excel)

# --- Contenido para REPORTE EMPALMERIA ---
opciones = ["DIEGO ARMANDO CHATEZ MARTINEZ","HAROLD ANDRES TORRES TEPUD","VICTOR ANDRES BOTINA ALVAREZ","CARLOS ANDRES MARCILLO","OMAR ALEXANDER DULCE LOPEZ",
            "YESID ALFONSO SANCHEZ DIAZ","ALDIVEY QUINAYAS MUÑOZ","DANIEL EDUARDO TROCHEZ MUÑOZ","ANDRES CAMILO ALEGRIA ALEGRIA","VICTOR ALIRIO ARDILA CELIS",
            "NASPIRAN ROSERO SEGUNDO JUBENAL","MARINO SANCHEZ GARCIA","DIEGO ARMANDO MUÑOZ SAAVEDRA","DIEGO ALEJANDRO VEGA GALEANO","RUTBEL TRUJILLO","VICTOR ALFONSO MORA"
            ]
marca = ["FICOTEL","CONDUMEX","HOME","FURUKAGUA","OFS","PLENUM"
            ]
capacidad = ["12","24","48","96","144","288"
            ]

if hoja_seleccionada == "REPORTE EMPALMERIA":
    st.subheader("Reporte Empalmería")
    with st.form("formulario_general"):
        st.subheader("Datos Generales")
        st.text_input("CLIENTE:", key="cliente", value=st.session_state.general_data.get('cliente', ''))
        st.text_input("DIRECCIÓN CLIENTE:", key="direccion_cliente", value=st.session_state.general_data.get('direccion_cliente', ''))
        #st.text_input("EMPALMADOR:", key="empalmador", value=st.session_state.general_data.get('empalmador', ''))
        st.selectbox("EMPALMADOR:", opciones, key="empalmador")
        st.text_input("CAMBIO, TICKET, OT:", key="ot", value=st.session_state.general_data.get('ot', ''))
        st.date_input("FECHA:", key="fecha", value=st.session_state.general_data.get('fecha', None))
        st.text_area("OBSERVACIONES:", key="observaciones", value=st.session_state.general_data.get('observaciones', ''))
        if st.form_submit_button("Guardar Datos Generales"):
            st.session_state.general_data = {
                'cliente': st.session_state.cliente, 'direccion_cliente': st.session_state.direccion_cliente,
                'empalmador': st.session_state.empalmador, 'ot': st.session_state.ot,
                'fecha': st.session_state.fecha, 'observaciones': st.session_state.observaciones
            }
            st.success("Datos generales guardados. Puedes cambiar de hoja o seguir diligenciando.")

    st.write("---")
    st.subheader("Agregar/Modificar Empalmes")
    current_empalme = {}
    if st.session_state.editando_empalme_idx is not None:
        current_empalme = st.session_state.empalmes_data[st.session_state.editando_empalme_idx]

    with st.form("formulario_empalmes"):
        st.text_input("# EMPALME:", key="num_empalme_current", value=current_empalme.get('num_empalme', ''))
        st.text_input("TIPO:", key="tipo_current", value=current_empalme.get('tipo', ''))
        st.text_input("DIRECCIÓN EMPALME:", key="direccion_empalme_current", value=current_empalme.get('direccion_empalme', ''))
        st.write("**CABLE IN**")
        cols_in = st.columns(2)
        cols_in[0].text_input("# INVENTARIO IN:", key="num_inventario_in_current", value=current_empalme.get('num_inventario_in', ''))
        cols_in[1].text_input("MARCA IN:", key="marca_in_current", value=current_empalme.get('marca_in', ''))
        cols_in[0].text_input("CAPACIDAD IN:", key="capacidad_in_current", value=current_empalme.get('capacidad_in', ''))
        cols_in[1].text_input("HILOS IN:", key="hilos_in_current", value=current_empalme.get('hilos_in', ''))
        st.write("**CABLE OUT**")
        cols_out = st.columns(2)
        cols_out[0].text_input("# INVENTARIO OUT:", key="num_inventario_out_current", value=current_empalme.get('num_inventario_out', ''))
        #cols_out[1].text_input("MARCA OUT:", key="marca_out_current", value=current_empalme.get('marca_out', ''))
        cols_out[0].text_input("CAPACIDAD OUT:", key="capacidad_out_current", value=current_empalme.get('capacidad_out', ''))
        cols_out[1].text_input("HILOS OUT:", key="hilos_out_current", value=current_empalme.get('hilos_out', ''))
        st.text_input("DISTANCIA:", key="distancia_current", value=current_empalme.get('distancia', ''))
        if st.form_submit_button("Guardar Empalme"):
            nuevo_empalme = {
                'num_empalme': st.session_state.num_empalme_current, 'tipo': st.session_state.tipo_current,
                'direccion_empalme': st.session_state.direccion_empalme_current,
                'num_inventario_in': st.session_state.num_inventario_in_current, 'marca_in': st.session_state.marca_in_current,
                'capacidad_in': st.session_state.capacidad_in_current, 'hilos_in': st.session_state.hilos_in_current,
                'num_inventario_out': st.session_state.num_inventario_out_current, 'marca_out': st.session_state.marca_out_current,
                'capacidad_out': st.session_state.capacidad_out_current, 'hilos_out': st.session_state.hilos_out_current,
                'distancia': st.session_state.distancia_current
            }
            if st.session_state.editando_empalme_idx is not None:
                st.session_state.empalmes_data[st.session_state.editando_empalme_idx] = nuevo_empalme
                st.success(f"Empalme #{st.session_state.editando_empalme_idx + 1} modificado exitosamente.")
                st.session_state.editando_empalme_idx = None
            else:
                st.session_state.empalmes_data.append(nuevo_empalme)
                st.success("Empalme agregado a la lista.")
            st.rerun()

    if st.session_state.empalmes_data:
        st.write("---")
        st.subheader("Lista de Empalmes Agregados")
        df_empalmes = pd.DataFrame(st.session_state.empalmes_data)
        st.dataframe(df_empalmes)
        
        col_mod_del = st.columns(2)
        with col_mod_del[0]:
            empalme_a_modificar = st.selectbox("Selecciona un empalme para modificar:", df_empalmes.index, format_func=lambda i: f"Empalme #{i+1}")
            if st.button("Cargar para Modificar"):
                st.session_state.editando_empalme_idx = empalme_a_modificar
                st.rerun()
        with col_mod_del[1]:
            empalme_a_eliminar = st.selectbox("Selecciona un empalme para eliminar:", df_empalmes.index, format_func=lambda i: f"Empalme #{i+1}")
            if st.button("Eliminar Empalme"):
                del st.session_state.empalmes_data[empalme_a_eliminar]
                st.success("Empalme eliminado.")
                st.rerun()

# --- Contenido para MATERIAL CONSUMIDO ---
elif hoja_seleccionada == "MATERIAL CONSUMIDO":
    st.subheader("Diligenciar Hoja: MATERIAL CONSUMIDO")
    
    df_materiales = cargar_listado_material_con_openpyxl(ruta_excel)
    
    if df_materiales is not None and not df_materiales.empty:
        opciones_materiales = [f"{row['Numero']} - {row['Descripcion']}" for index, row in df_materiales.iterrows()]
        tipos_material = ["MATERIAL INTERNO", "MATERIAL EXTERNO", "MATERIAL EMPALMERIA"]
        
        with st.form("form_material_consumido"):
            tipo_seleccionado = st.selectbox("Selecciona el tipo de material:", tipos_material)
            selected_material_str = st.selectbox("Busca y selecciona un material:", options=opciones_materiales)
            cantidad = st.number_input("Ingresa la cantidad:", min_value=1, step=1)
            
            if st.form_submit_button("Agregar material a la lista"):
                num_material_str = selected_material_str.split(' - ')[0]
                
                # Convertir la columna 'Numero' del DataFrame a un tipo numérico de forma segura
                df_materiales['Numero'] = pd.to_numeric(df_materiales['Numero'], errors='coerce')

                # Intentar convertir el valor extraído del selectbox a un tipo numérico
                try:
                    num_material = float(num_material_str)
                except ValueError:
                    st.error("El número de material no es un valor válido.")
                    num_material = None

                if num_material is not None:
                    # Usar .loc de forma más segura
                    item_seleccionado_df = df_materiales[df_materiales['Numero'] == num_material]
                    
                    if not item_seleccionado_df.empty:
                        item_seleccionado = item_seleccionado_df.iloc[0].to_dict()
                        item_seleccionado['Cantidad'] = cantidad
                        item_seleccionado['Tipo'] = tipo_seleccionado
                        st.session_state.materiales_consumidos.append(item_seleccionado)
                        st.success("Material agregado a la lista.")
                    else:
                        st.error(f"Material con número '{num_material_str}' no encontrado. Por favor, revisa el listado.")
                st.rerun()

    if st.session_state.materiales_consumidos:
        st.subheader("Materiales en la lista para diligenciar:")
        df_consumidos = pd.DataFrame(st.session_state.materiales_consumidos)
        st.dataframe(df_consumidos)

# --- Botón de descarga final (Unificado) ---
st.write("---")
st.header("Generar y Descargar Archivo Excel")

if st.button("Generar y Descargar Excel Completo"):
    try:
        libro = load_workbook(ruta_excel)

        # Diligenciar la hoja REPORTE EMPALMERIA
        if st.session_state.general_data or st.session_state.empalmes_data:
            diligenciar_hoja_empalmeria(libro, st.session_state.general_data, st.session_state.empalmes_data)

        # Diligenciar la hoja MATERIAL CONSUMIDO
        if st.session_state.materiales_consumidos:
            diligenciar_hoja_material_consumido(libro, st.session_state.materiales_consumidos)

        # Guardar el libro en memoria
        buffer = BytesIO()
        libro.save(buffer)
        buffer.seek(0)
        
        b64 = base64.b64encode(buffer.getvalue()).decode()
        href = f'<a href="data:application/octet-stream;base64,{b64}" download="formato_final_diligenciado.xlsx">Haz clic aquí para descargar el archivo.</a>'
        st.markdown(href, unsafe_allow_html=True)
        st.success("¡Archivo generado y listo para descargar!")

    except Exception as e:
        st.error(f"Ocurrió un error al generar el archivo: {e}")